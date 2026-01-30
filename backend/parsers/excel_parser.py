"""
Excel parser for xlsx and xls files.
Uses openpyxl for xlsx and xlrd for xls files.
"""

import io
from typing import Optional

from .base import BaseParser, ParseResult


class ExcelParser(BaseParser):
    """Parser for Excel bank statements (xlsx and xls formats)."""

    async def parse(
        self, file_content: bytes, filename: str, max_chars: int = 50000
    ) -> ParseResult:
        """
        Extract text from an Excel file.

        Args:
            file_content: Raw bytes of the Excel file
            filename: Original filename
            max_chars: Maximum characters to extract

        Returns:
            ParseResult with extracted text and metadata
        """
        ext = filename.lower().split(".")[-1]

        if ext == "xlsx":
            return await self._parse_xlsx(file_content, filename, max_chars)
        elif ext == "xls":
            return await self._parse_xls(file_content, filename, max_chars)
        else:
            return ParseResult(
                success=False,
                raw_text="",
                error=f"Дэмжигдээгүй Excel формат: {ext}",
            )

    async def _parse_xlsx(
        self, file_content: bytes, filename: str, max_chars: int
    ) -> ParseResult:
        """Parse xlsx files using openpyxl."""
        try:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(file_content), data_only=True)
            text_parts = []
            total_chars = 0
            total_rows = 0

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                text_parts.append(f"\n--- {sheet_name} ---\n")

                for row in sheet.iter_rows(values_only=True):
                    if total_chars >= max_chars:
                        text_parts.append("\n\n[Текст хэт урт тул товчилсон...]")
                        break

                    # Skip completely empty rows
                    if all(cell is None for cell in row):
                        continue

                    row_text = "\t".join(
                        str(cell).strip() if cell is not None else "" for cell in row
                    )
                    if row_text.strip():
                        text_parts.append(row_text)
                        total_chars += len(row_text)
                        total_rows += 1

                if total_chars >= max_chars:
                    break

            full_text = "\n".join(text_parts).strip()

            if not full_text:
                return ParseResult(
                    success=False,
                    raw_text="",
                    error="Excel файлаас өгөгдөл олдсонгүй.",
                )

            bank_name = self.detect_bank(full_text)

            return ParseResult(
                success=True,
                raw_text=full_text[:max_chars],
                metadata={
                    "sheets": workbook.sheetnames,
                    "rows": total_rows,
                    "bank_name": bank_name,
                    "format": "xlsx",
                    "filename": filename,
                },
            )

        except ImportError:
            return ParseResult(
                success=False,
                raw_text="",
                error="openpyxl санг суулгаагүй байна. pip install openpyxl",
            )
        except Exception as e:
            return ParseResult(
                success=False,
                raw_text="",
                error=f"Excel файл уншихад алдаа гарлаа: {str(e)}",
            )

    async def _parse_xls(
        self, file_content: bytes, filename: str, max_chars: int
    ) -> ParseResult:
        """Parse xls files using xlrd."""
        try:
            import xlrd

            workbook = xlrd.open_workbook(file_contents=file_content)
            text_parts = []
            total_chars = 0
            total_rows = 0

            for sheet_idx in range(workbook.nsheets):
                sheet = workbook.sheet_by_index(sheet_idx)
                text_parts.append(f"\n--- {sheet.name} ---\n")

                for row_idx in range(sheet.nrows):
                    if total_chars >= max_chars:
                        text_parts.append("\n\n[Текст хэт урт тул товчилсон...]")
                        break

                    row = sheet.row_values(row_idx)
                    # Skip completely empty rows
                    if all(cell == "" or cell is None for cell in row):
                        continue

                    row_text = "\t".join(
                        str(cell).strip() if cell else "" for cell in row
                    )
                    if row_text.strip():
                        text_parts.append(row_text)
                        total_chars += len(row_text)
                        total_rows += 1

                if total_chars >= max_chars:
                    break

            full_text = "\n".join(text_parts).strip()

            if not full_text:
                return ParseResult(
                    success=False,
                    raw_text="",
                    error="Excel файлаас өгөгдөл олдсонгүй.",
                )

            bank_name = self.detect_bank(full_text)

            return ParseResult(
                success=True,
                raw_text=full_text[:max_chars],
                metadata={
                    "sheets": [
                        workbook.sheet_by_index(i).name for i in range(workbook.nsheets)
                    ],
                    "rows": total_rows,
                    "bank_name": bank_name,
                    "format": "xls",
                    "filename": filename,
                },
            )

        except ImportError:
            return ParseResult(
                success=False,
                raw_text="",
                error="xlrd санг суулгаагүй байна. pip install xlrd",
            )
        except Exception as e:
            return ParseResult(
                success=False,
                raw_text="",
                error=f"Excel файл уншихад алдаа гарлаа: {str(e)}",
            )
